from pathlib import Path
import nltk
import pdfplumber
import docx
import openpyxl
import csv
import json
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
import markdown
import re


try:
    nltk.data.find("tokenizers/punkt")
except LookupError:
    nltk.download("punkt")

try:
    nltk.data.find("tokenizers/punkt_tab")
except LookupError:
    nltk.download("punkt_tab")


def load_text_files(data_dir="../data"):
    """Load .txt files"""
    docs = []
    base = Path(data_dir)
    
    for path in base.rglob("*.txt"):
        text = path.read_text(encoding="utf-8", errors="ignore")
        docs.append({
            "doc_id": path.stem,
            "text": text.strip(),
            "source": str(path),
            "type": "text"
        })
    return docs


def load_pdf_files(data_dir="../data"):
    """Load .pdf files"""
    docs = []
    base = Path(data_dir)
    
    for pdf_path in base.rglob("*.pdf"):
        text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            docs.append({
                "doc_id": pdf_path.stem,
                "text": text.strip(),
                "source": str(pdf_path),
                "type": "pdf"
            })
        except Exception as e:
            print(f"Error loading PDF {pdf_path}: {e}")
    return docs


def load_docx_files(data_dir="../data"):
    """Load .docx files (Microsoft Word)"""
    docs = []
    base = Path(data_dir)
    
    for docx_path in base.rglob("*.docx"):
        try:
            doc = docx.Document(docx_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            docs.append({
                "doc_id": docx_path.stem,
                "text": text.strip(),
                "source": str(docx_path),
                "type": "docx"
            })
        except Exception as e:
            print(f"Error loading DOCX {docx_path}: {e}")
    return docs


def load_excel_files(data_dir="../data"):
    """Load .xlsx, .xls files (Excel)"""
    docs = []
    base = Path(data_dir)
    
    for excel_path in base.rglob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            text_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}\n")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)
            
            docs.append({
                "doc_id": excel_path.stem,
                "text": "\n".join(text_parts),
                "source": str(excel_path),
                "type": "excel"
            })
        except Exception as e:
            print(f"Error loading Excel {excel_path}: {e}")
    return docs


def load_csv_files(data_dir="../data"):
    """Load .csv files"""
    docs = []
    base = Path(data_dir)
    
    for csv_path in base.rglob("*.csv"):
        try:
            text_parts = []
            with open(csv_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    row_text = "\t".join(row)
                    if row_text.strip():
                        text_parts.append(row_text)
            
            docs.append({
                "doc_id": csv_path.stem,
                "text": "\n".join(text_parts),
                "source": str(csv_path),
                "type": "csv"
            })
        except Exception as e:
            print(f"Error loading CSV {csv_path}: {e}")
    return docs

def load_json_files(data_dir="../data"):
    """Load .json files"""
    docs = []
    base = Path(data_dir)
    for json_path in base.rglob("*.json"):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                text = json.dumps(data, indent=2, ensure_ascii=False)
            
            docs.append({
                "doc_id": json_path.stem,
                "text": text,
                "source": str(json_path),
                "type": "json"
            })
        except Exception as e:
            print(f"Error loading JSON {json_path}: {e}")
    return docs


def load_html_files(data_dir="../data"):
    """Load .html, .htm files"""
    docs = []
    base = Path(data_dir)
    for html_path in base.rglob("*.html"):
        try:
            with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                for script in soup(["script", "style"]):
                    script.decompose()
                text = soup.get_text(separator='\n', strip=True)
            docs.append({
                "doc_id": html_path.stem,
                "text": text,
                "source": str(html_path),
                "type": "html"
            })
        except Exception as e:
            print(f"Error loading HTML {html_path}: {e}")
    return docs

def load_markdown_files(data_dir="../data"):
    """Load .md files"""
    docs = []
    base = Path(data_dir)
    for md_path in base.rglob("*.md"):
        try:
            text = md_path.read_text(encoding="utf-8", errors="ignore")
            html = markdown.markdown(text)
            soup = BeautifulSoup(html, 'html.parser')
            plain_text = soup.get_text(separator='\n', strip=True)
            docs.append({
                "doc_id": md_path.stem,
                "text": plain_text,
                "source": str(md_path),
                "type": "markdown"
            })
        except Exception as e:
            print(f"Error loading Markdown {md_path}: {e}")
    return docs


def load_xml_files(data_dir="../data"):
    """Load .xml files"""
    docs = []
    base = Path(data_dir)
    for xml_path in base.rglob("*.xml"):
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            def extract_text(element):
                text_parts = []
                if element.text:
                    text_parts.append(element.text.strip())
                for child in element:
                    text_parts.extend(extract_text(child))
                if element.tail:
                    text_parts.append(element.tail.strip())
                return text_parts
            text = "\n".join(filter(None, extract_text(root)))
            docs.append({
                "doc_id": xml_path.stem,
                "text": text,
                "source": str(xml_path),
                "type": "xml"
            })
        except Exception as e:
            print(f"Error loading XML {xml_path}: {e}")
    return docs


def load_all_documents(data_dir="../data"):
    """Load all supported file types"""
    all_docs = []
    loaders = [
        load_text_files,
        load_pdf_files,
        load_docx_files,
        load_excel_files,
        load_csv_files,
        load_json_files,
        load_html_files,
        load_markdown_files,
        load_xml_files
    ]
    for loader in loaders:
        try:
            docs = loader(data_dir)
            all_docs.extend(docs)
            print(f"Loaded {len(docs)} documents using {loader.__name__}")
        except Exception as e:
            print(f"Error in {loader.__name__}: {e}")
    return all_docs


def build_sentence_chunks(docs):
    chunks = []
    for doc in docs:
        sentences = nltk.sent_tokenize(doc["text"])
        for chunk_id, sentence in enumerate(sentences):
            cleaned = sentence.strip()
            if cleaned:
                chunks.append({
                    "doc_id": doc["doc_id"],
                    "chunk_id": chunk_id,
                    "text": cleaned,
                    "source": doc["source"],
                })
    return chunks


if __name__ == "__main__":
    docs = load_all_documents("../data")
    print(f"\nTotal documents loaded: {len(docs)}")
    chunks = build_sentence_chunks(docs)
    print(f"Total chunks created: {len(chunks)}")
    for i in chunks:
        print(i)
    if chunks:
        print("\nSample chunk:")
        print(chunks[0])